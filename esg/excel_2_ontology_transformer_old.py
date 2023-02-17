import openpyxl
from rdflib import Graph, URIRef, RDF, OWL, RDFS, BNode, Literal, Namespace
from rdflib.compat import decodeStringEscape
from tqdm import tqdm

EM_FACTOR_LABEL_SUFFIX = 'emission level'
EM_FACTOR_IRI_SUFFIX = 'EmissionLevel'

level_1_dict = dict()
level_2_dict = dict()
level_3_dict = dict()
level_3_plus_1_dict = dict()
level_3_plus_2_dict = dict()

defra_namespace = Namespace('https://spec.edmc.org/esg/ontology/EMFACT/DefraEmissionFactors/')
has_uom = URIRef(str(defra_namespace) + 'hasUnitOfMeasure')
has_defra_name = URIRef(str(defra_namespace) + 'hasDefraName')
has_defra_ghg_equivalent = URIRef(str(defra_namespace) + 'hasDefraGHGEquivalent')
defra_ghg_equivalent = URIRef(str(defra_namespace) + 'GHGEquivalent')
defra_ghg_equivalent_base = URIRef(str(defra_namespace) + 'GHGEquivalentBase')
defra_ghg = URIRef(str(defra_namespace) + 'GHG')
has_defra_ghg_equivalent_base = URIRef(str(defra_namespace) + 'hasGHGEquivalentBase')
has_defra_ghg = URIRef(str(defra_namespace) + 'hasGHG')
has_defra_factor = URIRef(str(defra_namespace) + 'hasFactor')


def __create_iri_from_string(iri_value: str) -> URIRef:
    regular_iri_string = str(iri_value)
    regular_iri_string = regular_iri_string.replace(' ', '')
    regular_iri_string = regular_iri_string.replace('/', 'slash')
    regular_iri_string = decodeStringEscape(str(regular_iri_string))
    regular_iri_string = regular_iri_string.replace('<', 'isSmallerThen')
    regular_iri_string = regular_iri_string.replace('>', 'isGreaterThen')
    regular_iri_string = str(defra_namespace) + regular_iri_string
    iri = URIRef(regular_iri_string)
    return iri


def __get_levelled_label(regular_label: str, column: int) -> Literal:
    levelled_label = Literal(' '.join([regular_label, EM_FACTOR_LABEL_SUFFIX, str(column - 1)]))
    return levelled_label


def __get_levelled_iri(regular_iri_string: str, column: int) -> URIRef:
    levelled_iri = __create_iri_from_string(regular_iri_string+EM_FACTOR_IRI_SUFFIX+str(column - 1))
    return levelled_iri


ontology = Graph(bind_namespaces="core")
ontology.add((has_defra_name, RDF.type, OWL.AnnotationProperty))
ontology.add((defra_ghg_equivalent, RDF.type, OWL.Class))
ontology.add((defra_ghg_equivalent_base, RDF.type, OWL.Class))
ontology.add((defra_ghg, RDF.type, OWL.Class))
ontology.add((has_defra_ghg_equivalent, RDF.type, OWL.ObjectProperty))
ontology.add((has_defra_ghg_equivalent_base, RDF.type, OWL.ObjectProperty))
ontology.add((has_uom, RDF.type, OWL.ObjectProperty))
ontology.add((has_defra_ghg, RDF.type, OWL.ObjectProperty))
ontology.add((has_defra_factor, RDF.type, OWL.DatatypeProperty))

wb = openpyxl.load_workbook('DefraEmissionFactors.xlsx', data_only=True)

sheets = wb.sheetnames

for sheet in sheets:
    current_sheet = wb[sheet]
    for row in tqdm(range(1, current_sheet.max_row + 1)):
        for column in range(1, 12):
            cell = current_sheet.cell(row=row, column=column)
            cell_value = cell.value
            if column > 1:
                parent_cell = current_sheet.cell(row=row, column=column - 1)
                parent_cell_value = parent_cell.value
            else:
                parent_cell = None
                parent_cell_value = None
            if column > 2:
                grand_parent_cell = current_sheet.cell(row=row, column=column - 2)
                grand_parent_cell_value = grand_parent_cell.value
            else:
                grand_parent_cell = None
                grand_parent_cell_value = None
            if column > 3:
                great_grand_parent_cell = current_sheet.cell(row=row, column=column - 3)
                great_grand_parent_cell_value = great_grand_parent_cell.value
            else:
                great_grand_parent_cell = None
                great_grand_parent_cell_value = None
            if cell_value is not None:
                if column == 1:
                    scope_iri = __create_iri_from_string(cell_value)
                    ontology.add((scope_iri, RDF.type, OWL.Class))
                    ontology.add((scope_iri, RDFS.label, Literal(str(cell_value))))
                    ontology.add((scope_iri, defra_namespace.hasDefraName, Literal(cell_value)))
                if column == 2:
                    level_1_iri = __get_levelled_iri(cell_value, column=column)
                    ontology.add((level_1_iri, RDF.type, OWL.Class))
                    levelled_label = __get_levelled_label(regular_label=str(cell_value), column=column)
                    ontology.add((level_1_iri, RDFS.label, levelled_label))
                    ontology.add((level_1_iri, defra_namespace.hasDefraName, Literal(cell_value)))
                    level_1_dict[cell_value] = level_1_iri.split('/')[-1]
                    if parent_cell_value is not None:
                        parent_iri = __create_iri_from_string(parent_cell_value)
                        ontology.add((level_1_iri, RDFS.subClassOf, parent_iri))
                if column == 3:
                    parent_level_1_local_name = level_1_dict[parent_cell_value]
                    level_2_iri = __get_levelled_iri(parent_level_1_local_name + str(cell_value), column=column)
                    ontology.add((level_2_iri, RDF.type, OWL.Class))
                    levelled_label = __get_levelled_label(regular_label=str(cell_value), column=column)
                    ontology.add((level_2_iri, RDFS.label, levelled_label))
                    ontology.add((level_2_iri, defra_namespace.hasDefraName, Literal(cell_value)))
                    level_2_dict[cell_value] = level_2_iri.split('/')[-1]
                    parent_iri = __create_iri_from_string(level_1_dict[parent_cell_value])
                    ontology.add((level_2_iri, RDFS.subClassOf, parent_iri))
                if column == 4:
                    parent_level_2_local_name = level_2_dict[parent_cell_value]
                    level_3_iri = __get_levelled_iri(parent_level_2_local_name + str(cell_value), column=column)
                    ontology.add((level_3_iri, RDF.type, OWL.Class))
                    levelled_label = __get_levelled_label(regular_label=str(cell_value), column=column)
                    ontology.add((level_3_iri, RDFS.label, levelled_label))
                    ontology.add((level_3_iri, defra_namespace.hasDefraName, Literal(cell_value)))
                    level_3_dict[cell_value] = level_3_iri.split('/')[-1]
                    parent_iri = __create_iri_from_string(level_2_dict[parent_cell_value])
                    ontology.add((level_3_iri, RDFS.subClassOf, parent_iri))
                if column == 6:
                    if grand_parent_cell_value is not None:
                        parent_level_3_local_name = level_3_dict[grand_parent_cell_value]
                        level_3_plus_1_iri = __create_iri_from_string(parent_level_3_local_name + str(cell_value))
                        ontology.add((level_3_plus_1_iri, RDF.type, OWL.Class))
                        ontology.add((level_3_plus_1_iri, RDFS.label, Literal(str(grand_parent_cell_value + ' ' + cell_value))))
                        ontology.add((level_3_plus_1_iri, has_defra_name, Literal(str(cell_value))))
                        level_3_plus_1_dict[cell_value] = level_3_plus_1_iri.split('/')[-1]
                        grand_parent_iri = __create_iri_from_string(level_3_dict[grand_parent_cell_value])
                        ontology.add((level_3_plus_1_iri, RDFS.subClassOf, grand_parent_iri))
                if column == 7:
                    uom_iri = __create_iri_from_string(cell_value)
                    ontology.add((uom_iri, RDF.type, OWL.NamedIndividual))
                    ontology.add((uom_iri, RDFS.label, Literal(cell_value)))
                    if parent_cell_value is not None:
                        level_3_plus_1_parent_local_name = level_3_plus_1_dict[parent_cell_value]
                        level_3_plus_1_parent_iri = __create_iri_from_string(level_3_plus_1_parent_local_name)
                        level_3_plus_2_iri = __create_iri_from_string(str(level_3_plus_1_parent_local_name) + str(cell_value))
                        ontology.add((level_3_plus_2_iri, RDF.type, OWL.Class))
                        ontology.add((level_3_plus_2_iri, RDFS.label, Literal(parent_cell_value + ' measured in ' + cell_value)))
                        ontology.add((level_3_plus_2_iri, RDFS.subClassOf, level_3_plus_1_parent_iri))
                        ontology.add((level_3_plus_2_iri, has_defra_name, Literal(str(cell_value))))
                        owl_restriction = BNode()
                        ontology.add((owl_restriction, RDF.type, OWL.Restriction))
                        ontology.add((owl_restriction, OWL.hasValue, uom_iri))
                        ontology.add((owl_restriction, OWL.onProperty, has_uom))
                        ontology.add((level_3_plus_2_iri, RDFS.subClassOf, owl_restriction))
                        level_3_plus_2_dict[str(cell_value)] = level_3_plus_2_iri.split('/')[-1]
                    else:
                        if great_grand_parent_cell_value is not None:
                            level_3_parent_local_name = level_3_dict[great_grand_parent_cell_value]
                            level_3_parent_iri = __create_iri_from_string(level_3_parent_local_name)
                            level_3_plus_2_iri = __create_iri_from_string(level_3_parent_local_name + str(cell_value))
                            ontology.add((level_3_plus_2_iri, RDF.type, OWL.Class))
                            ontology.add((level_3_plus_2_iri, RDFS.label, Literal(great_grand_parent_cell_value + ' measured in ' + cell_value)))
                            ontology.add((level_3_plus_2_iri, RDFS.subClassOf, level_3_parent_iri))
                            ontology.add((level_3_plus_2_iri, has_defra_name, Literal(str(cell_value))))
                            owl_restriction = BNode()
                            ontology.add((owl_restriction, RDF.type, OWL.Restriction))
                            ontology.add((owl_restriction, OWL.hasValue, uom_iri))
                            ontology.add((owl_restriction, OWL.onProperty, has_uom))
                            ontology.add((level_3_plus_2_iri, RDFS.subClassOf, owl_restriction))
                            level_3_plus_2_dict[str(cell_value)] = level_3_plus_2_iri.split('/')[-1]
                if column == 9:
                    child_cell = current_sheet.cell(row=row, column=column + 2)
                    child_cell_value = child_cell.value
                    equivalent_base_iri = __create_iri_from_string(cell_value)
                    equivalent_iri = __create_iri_from_string(str(cell_value)+'Equivalent')
                    if child_cell_value is not None:
                        equivalent_iri = __create_iri_from_string(str(cell_value)+str(child_cell_value)+'Equivalent')
                    ontology.add((equivalent_iri, RDF.type, OWL.NamedIndividual))
                    ontology.add((equivalent_iri, RDF.type, defra_ghg_equivalent))
                    ontology.add((equivalent_base_iri, RDF.type, OWL.NamedIndividual))
                    ontology.add((equivalent_base_iri, RDF.type, defra_ghg_equivalent_base))
                    if child_cell_value is not None:
                        if len(str(child_cell_value)) > 0:
                            ontology.add((equivalent_iri, RDFS.label, Literal(str(cell_value)+' '+str(child_cell_value)+' equivalent')))
                    else:
                        ontology.add((equivalent_iri, RDFS.label, Literal(cell_value)+' equivalent'))
                    ontology.add((equivalent_base_iri, RDFS.label, Literal(cell_value)))
                    ontology.add((equivalent_iri, has_defra_ghg_equivalent_base, equivalent_base_iri))
                    if child_cell_value is not None:
                        if len(str(child_cell_value)) > 0:
                            ontology.add((equivalent_iri, has_defra_factor, Literal(child_cell_value)))
                    
                    equivalent_string_components = str(cell_value).split(' ')
                    
                    equivalent_uom_string = equivalent_string_components[0]
                    equivalent_uom_iri = __create_iri_from_string(equivalent_uom_string)
                    ontology.add((equivalent_iri, has_uom, equivalent_uom_iri))

                    if len(equivalent_string_components) > 1:
                        if '(' not in str(cell_value).split(' ')[1]:
                            equivalent_gas_string = str(cell_value).split(' ')[1]
                            equivalent_gas_iri = __create_iri_from_string(equivalent_gas_string)
                            ontology.add((equivalent_gas_iri, RDF.type, OWL.NamedIndividual))
                            ontology.add((equivalent_gas_iri, RDF.type, defra_ghg))
                            ontology.add((equivalent_iri, has_defra_ghg, equivalent_gas_iri))
                    
                    level_3_plus_2_parent_cell_local_name = level_3_plus_2_dict[grand_parent_cell_value]
                    level_3_plus_2_parent_iri = __create_iri_from_string(level_3_plus_2_parent_cell_local_name)
                    owl_restriction = BNode()
                    ontology.add((owl_restriction, RDF.type, OWL.Restriction))
                    ontology.add((owl_restriction, OWL.hasValue, equivalent_iri))
                    ontology.add((owl_restriction, OWL.onProperty, has_defra_ghg_equivalent))
                    ontology.add((level_3_plus_2_parent_iri, RDFS.subClassOf, owl_restriction))
ontology.serialize('Defra.ttl')
ontology.commit()
